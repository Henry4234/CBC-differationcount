import tkinter as tk
import customtkinter  as ctk
from tksheet import Sheet
from tkinter import ttk
from tkinter import RIDGE, DoubleVar, StringVar, ttk, messagebox,IntVar, filedialog

from turtle import width
from PIL import Image
import os,pyodbc,openpyxl
import pandas as pd
import numpy as np

from sqlalchemy import create_engine
from sqlalchemy.engine import URL
import sqlalchemy as sa

from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl.formatting.rule import Rule,CellIsRule,FormulaRule


#SQL連線設定
connection_string = """DRIVER={ODBC Driver 17 for SQL Server};SERVER=localhost;DATABASE=bloodtest;UID=sa;PWD=1234"""
try:
    coxn = pyodbc.connect(connection_string)
except pyodbc.OperationalError:
    connection_string = "DRIVER={ODBC Driver 17 for SQL Server};SERVER=10.30.47.9;DATABASE=bloodtest;UID=henry423;PWD=1234"
finally:
    coxn = pyodbc.connect(connection_string)
    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": connection_string})
    engine = create_engine(connection_url)
##讀取95 reference
WB_95 = load_workbook("testdata\95.xlsx")
ws_95 = WB_95.worksheets[0]

##外邊框設置
#ex: A3:D5 要外粗邊框
# A3:A5的left style = "medium"
# D3:D5的right style = "medium"
# A3:D3的top style = "medium"
# A5:D5的bottom style = "medium"
def outter_border(worksheet,s_column, s_index, e_column , e_index,border_style):
    L_border = Border(left=Side(style=border_style),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    R_border = Border(left=Side(style="thin"),right=Side(style=border_style),top=Side(style="thin"),bottom=Side(style="thin"))
    T_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style=border_style),bottom=Side(style="thin"))
    B_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style=border_style))
    TL_border = Border(left=Side(style=border_style),right=Side(style="thin"),top=Side(style=border_style),bottom=Side(style="thin"))
    TR_border = Border(left=Side(style="thin"),right=Side(style=border_style),top=Side(style=border_style),bottom=Side(style="thin"))
    BL_border = Border(left=Side(style=border_style),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style=border_style))
    BR_border = Border(left=Side(style="thin"),right=Side(style=border_style),top=Side(style="thin"),bottom=Side(style=border_style))
    # 轉換欄位字母為對應的索引數字
    s_column_index = column_index_from_string(s_column)
    e_column_index = column_index_from_string(e_column)
    for row in range(s_index, e_index + 1):
        # 設定左邊框
        worksheet.cell(row=row, column=s_column_index).border = L_border
        # 設定右邊框
        worksheet.cell(row=row, column=e_column_index).border = R_border
    for col in range(s_column_index, e_column_index + 1):
        # 設定上邊框
        worksheet.cell(row=s_index, column=col).border = T_border
        # 設定下邊框
        worksheet.cell(row=e_index, column=col).border = B_border
    ##修正左上左下游上右下四角問題
    worksheet.cell(row=s_index, column=s_column_index).border = TL_border
    worksheet.cell(row=s_index, column=e_column_index).border = TR_border
    worksheet.cell(row=e_index, column=s_column_index).border = BL_border
    worksheet.cell(row=e_index, column=e_column_index).border = BR_border



class OUTPUT_SCORE:

    def __init__(self,master,oldmaster=None):
        self.master = master
        super().__init__()
        self.master.protocol("WM_DELETE_WINDOW",lambda: self.back(oldmaster))        
        ctk.set_default_color_theme("blue")
        self.master.title("成績匯出")
        self.master.geometry("1200x650")
        self.master.config(background='#FFEEDD')
        self.master.grid_rowconfigure(0, weight=1)
        self.master.grid_columnconfigure(1, weight=2)
        #框架設計
        self.labelframe_1 = ctk.CTkFrame(self.master,corner_radius=0,fg_color="#FFDCB9",bg_color="#FFEEDD")
        self.labelframe_1.grid(row=0, column=0, rowspan=2,sticky='nsew')
        self.labelframe_2 = ctk.CTkFrame(self.master,height=1200,fg_color="#FFEEDD",bg_color="#FFEEDD")
        self.labelframe_2.grid(row=0, column=1,sticky='nsew')
        self.testicon = ctk.CTkImage(Image.open("assets\output.png"),size=(150,150))
        #匯入標題
        self.welcome = ctk.CTkLabel(
                    self.labelframe_1, 
                    image=self.testicon,
                    compound="top",
                    text = "成績匯出", 
                    # fg_color='#FFDCB9',
                    font=('微軟正黑體',36),
                    text_color="#000000",
                    width=120,height=120
                    )
        self.welcome.grid(row=0,column=0,padx=10)
        self.all_combobox = ctk.CTkComboBox(
                    self.labelframe_1,
                    command=self.callback_all,
                    fg_color='#FFDCB9',
                    button_color='#FF9900',
                    values=["","All","台北","基隆","林口","桃園","雲林","嘉義","高雄","土城","大里仁愛"],
                    font=('微軟正黑體',22),
                    text_color="#000000",
                    width=120,height=50
                    )
        self.all_combobox.grid(row=1,column=0,padx=20, pady=20,sticky='ew')
        self.year_combobox = ctk.CTkComboBox(
                    self.labelframe_1, 
                    fg_color='#FFDCB9',
                    button_color='#FF9900',
                    values=[""],
                    font=('微軟正黑體',22),
                    text_color="#000000",
                    width=120,height=50
                    )
        self.year_combobox.grid(row=2,column=0,padx=20, pady=20,sticky='nsew')
        self.btn_search = ctk.CTkButton(
            self.labelframe_1,
            text="搜尋",
            anchor=tk.CENTER,
            command=self.search_gai,
            fg_color="#FF9224",
            bg_color='#FFDCB9',
            font=('微軟正黑體',22,'bold'),
            text_color="#000000",
            width=120,height=20,
        )
        self.btn_search.grid(row=3,column=0,padx=20,pady=30,sticky='ew')
        #清除資料
        self.btn_clear = ctk.CTkButton(
            self.labelframe_1,
            text="清除搜尋結果",
            anchor=tk.CENTER,
            command=self.clear,
            fg_color="#FF9224",
            bg_color='#FFEEDD',
            font=('微軟正黑體',22,'bold'),
            text_color="#000000",
            width=120,
        )
        self.btn_clear.grid(row=4,column=0,padx=20,sticky='ew')
        self.label_testtable=ctk.CTkLabel(
            self.labelframe_2,
            fg_color="#FFEEDD",
            bg_color="#FFEEDD",
            text="預覽結果",
            font=('微軟正黑體',26,'bold'),
            text_color="#000000"
            # height=300,width=950,
            )
        self.label_testtable.grid(row=0,column=0,columnspan=3,sticky='s')
        self.frame_testtable=ctk.CTkFrame(
            self.labelframe_2,
            fg_color="#FFEEDD",
            bg_color="#FFEEDD",
            height=300,width=950,
            )
        self.frame_testtable.grid(row=1,column=0,columnspan=3,padx=30,pady=5)
        ##tksheet_1結果表格設計
        self.sht_result = Sheet(
            self.frame_testtable,
            column_width=185,
            width=950,height=300,
        )
        self.sht_result.headers(["院區","考生姓名","考片_ID","考試次數","交卷時間"])
        self.sht_result.enable_bindings(('single_select',"row_select","drag_select"))
        
        self.sht_result.grid(row=0,column=0,sticky='ew')
        #返回主介面
        self.btn_back = ctk.CTkButton(
            self.labelframe_1,
            text="返回",
            anchor=tk.CENTER,
            command=lambda: self.back(oldmaster),
            fg_color="#FF9224",
            bg_color='#FFEEDD',
            font=('微軟正黑體',22,'bold'),
            text_color="#000000",
            width=120,
        )
        self.btn_back.grid(row=5,column=0,padx=20,pady=80,sticky='ew')
        self.btn_export = ctk.CTkButton(
            self.labelframe_2,
            text="匯出考試結果",
            anchor=tk.CENTER,
            command= self.output_testresult,
            fg_color="#FF9224",
            bg_color='#FFEEDD',
            font=('微軟正黑體',22,'bold'),
            text_color="#000000",
            width=160,height=90,
            state='disabled',
        )
        self.btn_export.grid(row=2,column=0,padx=20,pady=20)
        self.btn_rawexport = ctk.CTkButton(
            self.labelframe_2,
            text="匯出考生原始結果",
            anchor=tk.CENTER,
            command=self.output_percent,
            fg_color="#FF9224",
            bg_color='#FFEEDD',
            font=('微軟正黑體',22,'bold'),
            text_color="#000000",
            width=160,height=90,
            state='disabled',
        )
        self.btn_rawexport.grid(row=2,column=1,padx=20,pady=20)
        self.btn_sasexport = ctk.CTkButton(
            self.labelframe_2,
            text="匯出SAS",
            anchor=tk.CENTER,
            command=self.output_sas,
            fg_color="#FF9224",
            bg_color='#FFEEDD',
            font=('微軟正黑體',22,'bold'),
            text_color="#000000",
            width=160,height=90,
            state='disabled',
        )
        self.btn_sasexport.grid(row=2,column=2,padx=20,pady=20)
        ##字型標準檔
        self.ft1 = openpyxl.styles.Font(name='Times New Roman', size=14)  #內文字體設定
        self.ft2 = openpyxl.styles.Font(name='Times New Roman', size=12)  #內文字體設定
        ##框線標準檔
        ##框線設定
        left, right, top, bottom = [Side(style='thin',color='000000')]*4  #新增框線
        T_left, T_right, T_top, T_bottom = [Side(style='medium',color='000000')]*4  #新增框線
        self.bd1 = Border(left=T_left, right=T_right, top=T_top, bottom=T_bottom)
        self.bd2 = Border(left=left, right=right, top=top, bottom=bottom)
    #如果第一個選單式"ALL"的話，應該要把年分的欄位清空，變成不可選填。
    def callback_all(self,event):
        h_site = self.all_combobox.get()
        self.btn_export.configure(state='disabled')
        self.btn_rawexport.configure(state='disabled')
        self.btn_sasexport.configure(state='disabled')
        self.sht_result.set_sheet_data([[]])
        if h_site == "All":
            self.year_combobox.set("")
            self.year_combobox.configure(values = [])
            self.year_combobox.configure(state = 'disable')
        else:
            self.year_combobox.configure(state = 'normal')
            self.year_combobox.set("")
            event_year = """SELECT DISTINCT [bloodinfo].[year] 
FROM [bloodtest].[dbo].[blood_final] 
JOIN [bloodtest].[dbo].[id] 
ON [id].[No] = [blood_final].[test_id]
JOIN [bloodtest].[dbo].[bloodinfo] 
ON [bloodinfo].[smear_id] = [blood_final].[smear_id]
WHERE [id].[院區]='%s';;"""%(h_site)
            with coxn.cursor() as cursor:
                cursor.execute(event_year)
                un_year = cursor.fetchall()
            un_year = [e[0] for e in un_year]
            str_un_year =[]
            for item in un_year:
                item  = str(item)
                str_un_year.append(item)
            if len(str_un_year) == 0:
                self.year_combobox.set("")
                self.year_combobox.configure(values = [])
            self.year_combobox.configure(values = str_un_year)
##################↓↓↓↓↓搜尋: 查看當前所選擇的院區以及年分↓↓↓↓↓##################
    def search_gai(self):
        h_site = self.all_combobox.get()
        testyear = self.year_combobox.get()
        if h_site !="":
            if h_site=="All":
                srh = """SELECT DISTINCT [id].[院區], [id].[ac],[blood_final].[smear_id],[blood_final].[count],CONVERT (varchar,[blood_final].[timestamp],100)
                FROM [bloodtest].[dbo].[blood_final] 
                JOIN [bloodtest].[dbo].[id] 
                ON [id].[No] = [blood_final].[test_id];"""
                self.btn_export.configure(state='normal')
                self.btn_rawexport.configure(state='normal')
                self.btn_sasexport.configure(state='normal')
            elif testyear !="":
                srh = """SELECT DISTINCT [id].[院區], [id].[ac],[blood_final].[smear_id],[blood_final].[count],CONVERT (varchar,[blood_final].[timestamp],100)
                    FROM [bloodtest].[dbo].[blood_final] 
                    JOIN [bloodtest].[dbo].[id] 
                    ON [id].[No] = [blood_final].[test_id]
                    JOIN [bloodtest].[dbo].[bloodinfo] 
                    ON [bloodinfo].[smear_id] = [blood_final].[smear_id]
                    WHERE [id].[院區] = '%s' AND [bloodinfo].[year] = %d ;"""%(h_site,int(testyear))
                self.btn_export.configure(state='normal')
                self.btn_rawexport.configure(state='normal')
                self.btn_sasexport.configure(state='normal')
            else:
                tk.messagebox.showerror(title='土城醫院檢驗科', message='請選擇年份!')
                return
        else:
            tk.messagebox.showerror(title='土城醫院檢驗科', message='請選擇院區!所有院區請選擇ALL')
            return
        with coxn.cursor() as cursor:
            cursor.execute(srh)
            data_sht_1 = cursor.fetchall()
        # print(data_sht_1)
        self.sht_result.set_sheet_data(data_sht_1,reset_col_positions=True,reset_row_positions=True)
        # return
#########################################↑↑↑↑↑搜尋: 查看當前所選擇的院區以及年分↑↑↑↑↑#########################################
#########################################↓↓↓↓↓匯出考試結果: 包含考片正確答案，↓↓↓↓↓#########################################
    def output_testresult(self):
        h_site = self.all_combobox.get()
        testyear = self.year_combobox.get()
        if h_site !="":
            if h_site=="All":
                get_ans_bool = """SELECT [bloodinfo_ans2].[smear_id],[bloodinfo_ans2].[celltype],[bloodinfo_ans2].[must],[bloodinfo_ans2].[mustnot]
FROM [bloodtest].[dbo].[bloodinfo_ans2]
WHERE [bloodinfo_ans2].[smear_id] IN (
	SELECT DISTINCT [blood_final].[smear_id] FROM [bloodtest].[dbo].[blood_final]
	);"""
                get_data_matrix = """SELECT [id].[院區],[id].[ac],[bloodinfo].[year],[blood_final].[smear_id],[blood_final].[count],[blood_final].[celltype],[blood_final].[matrix_value],[blood_final].[timestamp]
FROM [bloodtest].[dbo].[blood_final]
JOIN [bloodtest].[dbo].[id] ON [id].[No] = [blood_final].[test_id]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id] = [blood_final].[smear_id];"""
                get_final = """SELECT DISTINCT [id].[院區],[id].[ac],[bloodinfo].[year],[blood_final].[smear_id],[blood_final].[count],[test_data].[must],[test_data].[mustnot],[test_data].[abn_lym],[test_data].[score]
FROM [blood_final]
JOIN [test_data] 
ON [test_data].[test_id]=[blood_final].[test_id] 
AND [test_data].[smear_id]=[blood_final].[smear_id] 
AND [test_data].[count]=[blood_final].[count]
JOIN [bloodtest].[dbo].[id] ON [id].[No] = [test_data].[test_id]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id] = [test_data].[smear_id];"""
                file_name = "All_"
            elif testyear !="":
                get_ans_bool = """SELECT [bloodinfo_ans2].[smear_id],[bloodinfo_ans2].[celltype],[bloodinfo_ans2].[must],[bloodinfo_ans2].[mustnot]
FROM [bloodtest].[dbo].[bloodinfo_ans2]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id]= [bloodinfo_ans2].[smear_id]
WHERE [bloodinfo_ans2].[smear_id] IN (
	SELECT DISTINCT [blood_final].[smear_id] 
	FROM [bloodtest].[dbo].[blood_final]
	JOIN [bloodtest].[dbo].[id] ON [id].[No] = [blood_final].[test_id]
	WHERE [id].[院區] = '%s')
AND [bloodinfo].[year]=%d"""%(h_site,int(testyear))
                get_data_matrix = """SELECT [id].[院區],[id].[ac],[bloodinfo].[year],[blood_final].[smear_id],[blood_final].[count],[blood_final].[celltype],[blood_final].[matrix_value],[blood_final].[timestamp]
FROM [bloodtest].[dbo].[blood_final]
JOIN [bloodtest].[dbo].[id] ON [id].[No] = [blood_final].[test_id]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id] = [blood_final].[smear_id]
WHERE [id].[院區] = '%s' AND [bloodinfo].[year]=%d;"""%(h_site,int(testyear))
                get_final = """SELECT DISTINCT [id].[院區],[id].[ac],[bloodinfo].[year],[blood_final].[smear_id],[blood_final].[count],[test_data].[must],[test_data].[mustnot],[test_data].[abn_lym],[test_data].[score]
FROM [blood_final]
JOIN [test_data] 
ON [test_data].[test_id]=[blood_final].[test_id] 
AND [test_data].[smear_id]=[blood_final].[smear_id] 
AND [test_data].[count]=[blood_final].[count]
JOIN [bloodtest].[dbo].[id] ON [id].[No] = [test_data].[test_id]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id] = [test_data].[smear_id]
WHERE [id].[院區]='%s' AND [bloodinfo].[year]=%d;"""%(h_site,int(testyear))
                file_name = "%d_%s_"%(int(testyear),h_site)
            else:
                tk.messagebox.showerror(title='土城醫院檢驗科', message='請選擇年份!')
                return
        else:
            tk.messagebox.showerror(title='土城醫院檢驗科', message='請選擇院區!所有院區請選擇ALL')
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet1"
        
        title = ["考片ID","must/mustnot","plasma cell","abnormal lympho","megakaryocyte","nRBC","blast","metamyelocyte","eosinophil","plasmacytoid","promonocyte","promyelocyte","band neutropil","basopil","atypical lymphocyte","hypersegmented neutrophil","myelocyte","segmented neutrophil","lymphocyte","monocyte"]
        ws.append(title)
        with engine.begin() as conn:
            ans_bool_data = pd.read_sql_query(sa.text(get_ans_bool), conn)
        lst = ans_bool_data['smear_id'].unique()

        n=2
        for i in range(0,len(lst)):
            flt = (ans_bool_data['smear_id']==lst[i])
            flt_dic = ans_bool_data.loc[flt]
            dic_must = flt_dic.set_index('celltype')['must'].to_dict()
            dic_mustnot = flt_dic.set_index('celltype')['mustnot'].to_dict()
            # ws.append([item])
            dic_must['smear_id'] = lst[i]
            dic_mustnot['smear_id'] = lst[i]
            ##加入解答標題列(考片ID、must/mustnot欄位)
            ws.cell(row=n, column=1, value=lst[i])  #考片ID輸入
            ws.cell(row=n+1, column=1, value=lst[i])    #考片ID輸入
            ws.cell(row=n, column=2, value='must')  #分別放入must/mustnot
            ws.cell(row=n+1, column=2, value='mustnot') #分別放入must/mustnot
            ws[get_column_letter(2) + str(n)].font = self.ft2   ##放入字型設定
            ws[get_column_letter(2) + str(n+1)].font = self.ft2 ##放入字型設定
            ###加入解答標題列(因為原本應該是直的，需要利用dictionary置轉)
            for col in range(3, 21):  # 从第二列开始遍历
                celltype = ws.cell(row=1, column=col).value
                if celltype in dic_must or dic_mustnot:
                    value_must = dic_must[celltype]
                    value_mustnot = dic_mustnot[celltype]
                    ws.cell(row=n, column=col, value=value_must)
                    ws.cell(row=n+1, column=col, value=value_mustnot)
                    ws[get_column_letter(col) + str(n)].font = self.ft2
                    ws[get_column_letter(col) + str(n+1)].font = self.ft2
            n+=2
        ##加入考生標題列
        title_2=['院區','姓名','year','考片ID','count','celltype','matrix_value','timestamp']
        for i in range(0,len(title_2)): #第二個表格標題
            char = get_column_letter(i+1)
            ws[char + str(len(lst)*2 + 4)] = title_2[i]
        with engine.begin() as conn:
            matrix_data = pd.read_sql_query(sa.text(get_data_matrix), conn)
        #加入data
        for j in range(0,len(matrix_data)):
            ws.append(matrix_data.iloc[j].to_list())
        ##加入考生結果標題列
        title_3=['院區','姓名','year','考片ID','count','must','mustnot','abn_lym','score']
        for k in range(0,len(title_3)):
            char = get_column_letter(k+1)
            ws[char + str(len(lst)*2 + 6 + len(matrix_data))] = title_3[k]
        #加入data
        with engine.begin() as conn:
            final_data = pd.read_sql_query(sa.text(get_final), conn)
        for i in range(0,len(final_data)):
            ws.append(final_data.iloc[i].to_list())
        ##########################格式設定##########################
        # 建立填充和字體的樣式
        fill = PatternFill(start_color="FFD2D2", end_color="FFD2D2", fill_type="solid")
        red_font = Font(color="FF0000")
        # 建立條件式格式設定規則
        rule_1 = CellIsRule(operator="equal", formula=["TRUE"], stopIfTrue=False, fill=fill, font=red_font)
        ##調整第一個table(解答boolin值)
        for col in range(1,21):
            char = get_column_letter(col)   #轉換字串
            ws[char + str(1)].font = self.ft1   #修改標題字型
            ws[char + str(1)].alignment = Alignment(horizontal='center',vertical='center')  #修改標題置中
            for row in range(1,len(lst)*2 + 2): #迴圈讓第一個table的內容新建框線
                ws[char + str(row)].border = self.bd2
            ws.column_dimensions[get_column_letter(col)].auto_size = True   #設定自動調整欄寬
        ##加入條件格式設定
        formatting_range = "C2:T%d"%(len(lst)*2 + 1)
        ws.conditional_formatting.add(formatting_range,rule_1)
        ##調整第二個table(考生結果)
        for col in range(1,9):
            char = get_column_letter(col)   #轉換字串
            ws[char + str(len(lst)*2 + 4)].font = self.ft1   #修改標題字型
            ws[char + str(len(lst)*2 + 4)].alignment = Alignment(horizontal='center',vertical='center')  #修改標題置中
            for row in range(len(lst)*2 + 4, len(lst)*2 + 5 + len(matrix_data)):    #迴圈讓第二個table的內容修改成內容字型
                ws[char + str(row)].border = self.bd2
        ##調整第三個table(考生三個boolin值，一定要打到，不可打到細胞，abn_lym)
        for col in range(1,10):
            char = get_column_letter(col)   #轉換字串
            ws[char + str(len(lst)*2 + 6 + len(matrix_data))].font = self.ft1   #修改標題字型
            ws[char + str(len(lst)*2 + 6 + len(matrix_data))].alignment = Alignment(horizontal='center',vertical='center')  #修改標題置中
            for row in range(len(lst)*2 + 6 + len(matrix_data),len(lst)*2 + 7 + len(matrix_data)+len(final_data)):    #迴圈讓第三個table的內容修改成內容字型
                ws[char + str(row)].border = self.bd2
        ##調整外框
        ###table_1
        outter_border(ws,"A", 1, "T" , len(lst)*2 + 1,"medium")
        ###table_2
        outter_border(ws,"A", len(lst)*2 + 4 , "H" , len(lst)*2 + 4 + len(matrix_data),"medium")
        ###table_3
        outter_border(ws,"A", len(lst)*2 + 6 + len(matrix_data), "I" , len(lst)*2 + 6 + len(matrix_data)+len(final_data),"medium")


        wb.save(file_name + "考生結果.xlsx")
        filepath = ".//%s考生結果.xlsx"%(file_name)
        if os.path.isfile(filepath):
            tk.messagebox.showinfo(title='土城醫院檢驗科', message='檔案新增成功!')
        else:
            tk.messagebox.showerror(title='土城醫院檢驗科', message='檔案新增失敗QQ')
#########################################↑↑↑↑↑匯出考試結果: 包含考片正確答案，↑↑↑↑↑#########################################
##################↓↓↓↓↓匯出考生原始結果: 包含考片正確答案、院區、姓名、考片ID、考試次數、百分比、上下限↓↓↓↓↓##################
    def output_percent(self):
        h_site = self.all_combobox.get()
        testyear = self.year_combobox.get()
        if h_site !="":
            if h_site=="All":
                get_ans = """SELECT [bloodinfo_ans2].[smear_id],[bloodinfo].[count_value],[bloodinfo_ans2].[celltype],[bloodinfo_ans2].[value]
FROM [bloodtest].[dbo].[bloodinfo_ans2]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id] = [bloodinfo_ans2].[smear_id]
WHERE [bloodinfo_ans2].[smear_id] IN (SELECT DISTINCT [blood_final].[smear_id] FROM [bloodtest].[dbo].[blood_final]);"""
                get_data = """SELECT [id].[院區],[id].[ac],[bloodinfo].[year],[blood_final].[smear_id],[blood_final].[count],[blood_final].[celltype],[blood_final].[percent_value],[blood_final].[timestamp]
FROM [bloodtest].[dbo].[blood_final]
JOIN [bloodtest].[dbo].[id] ON [id].[No] = [blood_final].[test_id]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id] = [blood_final].[smear_id];"""
                file_name = "All_"
            elif testyear !="":
                get_ans = """SELECT [bloodinfo_ans2].[smear_id],[bloodinfo].[count_value],[bloodinfo_ans2].[celltype],[bloodinfo_ans2].[value]
FROM [bloodtest].[dbo].[bloodinfo_ans2]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id]= [bloodinfo_ans2].[smear_id]
WHERE [bloodinfo_ans2].[smear_id] IN (
	SELECT DISTINCT [blood_final].[smear_id] 
	FROM [bloodtest].[dbo].[blood_final]
	JOIN [bloodtest].[dbo].[id] ON [id].[No] = [blood_final].[test_id]
	WHERE [id].[院區] = '%s')
AND [bloodinfo].[year]=%d;"""%(h_site,int(testyear))
                get_data = """SELECT [id].[院區],[id].[ac],[bloodinfo].[year],[blood_final].[smear_id],[blood_final].[count],[blood_final].[celltype],[blood_final].[percent_value],[blood_final].[timestamp]
FROM [bloodtest].[dbo].[blood_final]
JOIN [bloodtest].[dbo].[id] ON [id].[No] = [blood_final].[test_id]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id] = [blood_final].[smear_id]
WHERE [id].[院區] = '%s' AND [bloodinfo].[year]=%d;"""%(h_site,int(testyear))
                file_name = "%d_%s_"%(int(testyear),h_site)
            else:
                tk.messagebox.showerror(title='土城醫院檢驗科', message='請選擇年份!')
                return
        else:
            tk.messagebox.showerror(title='土城醫院檢驗科', message='請選擇院區!所有院區請選擇ALL')
            return
        with engine.begin() as conn:
            datadict2 = pd.read_sql_query(sa.text(get_ans), conn)
        lst = datadict2['smear_id'].unique()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet1"
        ##加入95 reference data
        wb.create_sheet('95')
        for row in ws_95.iter_rows(values_only=True):
            wb['95'].append(row)
        #加入解答標題列
        title = ["考片ID","count_value","plasma cell","abnormal lympho","megakaryocyte","nRBC","blast","metamyelocyte","eosinophil","plasmacytoid","promonocyte","promyelocyte","band neutropil","basopil","atypical lymphocyte","hypersegmented neutrophil","myelocyte","segmented neutrophil","lymphocyte","monocyte"]
        ws.append(title)
        n=2
        for i in range(0,len(lst)):
            flt = (datadict2['smear_id']==lst[i])   #建立filter，把相對應的考片ID數值取出來
            flt_dic = datadict2.loc[flt]    #取值
            dic_2 = flt_dic.set_index('celltype')['value'].to_dict()    #建立dictionary，結構長這樣 dict_2={'plasma cell':0,0,'abnormal lympho'=1.9}
            dic_2['smear_id'] = lst[i]  #在dict_2新增smear_id
            dic_2['count_value'] = flt_dic['count_value'].iloc[0]   #在dict_2新增count_value(需要數幾顆)
            ##加入
            ws.cell(row=n, column=1, value=lst[i])
            ws.cell(row=n, column=2, value=flt_dic['count_value'].iloc[0])
            ws[get_column_letter(1) + str(n)].font = self.ft2
            ws[get_column_letter(2) + str(n)].font = self.ft2
            # 逐列遍历 row(1)，查找对应的值并放入 row(n)
            for col in range(3, 21):  # 从第二列开始遍历
                celltype = ws.cell(row=1, column=col).value
                if celltype in dic_2:
                    value = dic_2[celltype]
                    ws.cell(row=n, column=col, value=value)
                    ws[get_column_letter(col)+str(n)].font = self.ft2
            n+=1
        ##加入考生標題列
        title_2=['院區','姓名','year','考片ID','count','celltype','percent_value','timestamp','lower','upper']
        for i in range(0,len(title_2)):
            char = get_column_letter(i+1)
            ws[char + str(len(lst)+4)] = title_2[i]
            ws[char + str(len(lst)+4)].font = self.ft1
        with engine.begin() as conn:
            data = pd.read_sql_query(sa.text(get_data), conn)
        #加入data
        for j in range(0,len(data)):
            ws.append(data.iloc[j].to_list())
            ws["I" + str(len(lst)+5+j)] = "=VLOOKUP(ROUND(HLOOKUP(F%s,$A$1:$T$%s,MATCH(D%s,$A$1:$A$%s,0),FALSE),0),'95'!$A$1:$K$103,IFS(VLOOKUP(D%s,$A$2:$B$%s,2,FALSE)=100,2,VLOOKUP(D%s,$A$2:$B$%s,2,FALSE)=200,4,VLOOKUP(D%s,$A$2:$B$%s,2,FALSE)=500,6),FALSE)"%(str(len(lst)+5+j),str(len(lst)+1),str(len(lst)+5+j),str(len(lst)+1),str(len(lst)+5+j),str(len(lst)+1),str(len(lst)+5+j),str(len(lst)+1),str(len(lst)+5+j),str(len(lst)+1))
            ws["J" + str(len(lst)+5+j)] = "=VLOOKUP(ROUND(HLOOKUP(F%s,$A$1:$T$%s,MATCH(D%s,$A$1:$A$%s,0),FALSE),0),'95'!$A$1:$K$103,IFS(VLOOKUP(D%s,$A$2:$B$%s,2,FALSE)=100,3,VLOOKUP(D%s,$A$2:$B$%s,2,FALSE)=200,5,VLOOKUP(D%s,$A$2:$B$%s,2,FALSE)=500,7),FALSE)"%(str(len(lst)+5+j),str(len(lst)+1),str(len(lst)+5+j),str(len(lst)+1),str(len(lst)+5+j),str(len(lst)+1),str(len(lst)+5+j),str(len(lst)+1),str(len(lst)+5+j),str(len(lst)+1))
        ##########################格式設定##########################
        ##調整第一個table(解答數值)
        for col in range(1,21):
            char = get_column_letter(col)   #轉換字串
            ws[char + str(1)].font = self.ft1   #修改標題字型
            ws[char + str(1)].alignment = Alignment(horizontal='center',vertical='center')  #修改標題置中
            for row in range(1,len(lst) + 2): #迴圈讓第一個table的內容新建框線
                ws[char + str(row)].border = self.bd2
            ws.column_dimensions[get_column_letter(col)].auto_size = True
        ##調整第二個table(考生結果)
        for col in range(1,11):
            char = get_column_letter(col)   #轉換字串
            ws[char + str(len(lst) + 4)].font = self.ft1   #修改標題字型
            ws[char + str(len(lst) + 4)].alignment = Alignment(horizontal='center',vertical='center')  #修改標題置中
            for row in range(len(lst) + 4, len(lst) + 5 + len(data)):    #迴圈讓第二個table的內容修改成內容字型
                ws[char + str(row)].border = self.bd2
        ##調整外框
        ###table_1
        outter_border(ws,"A", 1, "T" , len(lst) + 1,"medium")
        ###table_2
        outter_border(ws,"A", len(lst) + 4, "J" , len(lst) + 4 + len(data),"medium")

        wb.save(file_name + "考生原始成績.xlsx")
        filepath = ".//%s考生原始成績.xlsx"%(file_name)
        if os.path.isfile(filepath):
            tk.messagebox.showinfo(title='土城醫院檢驗科', message='檔案新增成功!')
        else:
            tk.messagebox.showerror(title='土城醫院檢驗科', message='檔案新增失敗QQ')
##################↑↑↑↑↑匯出考生原始結果: 包含考片正確答案、院區、姓名、考片ID、考試次數、百分比、上下限↑↑↑↑↑##################
##################↓↓↓↓↓匯出SAS↓↓↓↓↓##################
    def output_sas(self):
        h_site = self.all_combobox.get()
        testyear = self.year_combobox.get()
        if h_site !="":
            if h_site=="All":
                get_sas = """SELECT [id].[院區],[id].[ac],[bloodinfo].[year],[blood_final].[smear_id],[blood_final].[count],[blood_final].[celltype],
[blood_final].[matrix_value],[test_data].[must],[test_data].[mustnot],[test_data].[abn_lym],[test_data].[score]
FROM [blood_final]
JOIN [test_data] 
ON [test_data].[test_id]=[blood_final].[test_id] 
AND [test_data].[smear_id]=[blood_final].[smear_id] 
AND [test_data].[count]=[blood_final].[count]
JOIN [bloodtest].[dbo].[id] ON [id].[No] = [test_data].[test_id]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id] = [test_data].[smear_id];"""
                file_name = "All_"
            elif testyear !="":
                get_sas = """SELECT [id].[院區],[id].[ac],[bloodinfo].[year],[blood_final].[smear_id],[blood_final].[count],[blood_final].[celltype],
[blood_final].[matrix_value],[test_data].[must],[test_data].[mustnot],[test_data].[abn_lym],[test_data].[score]
FROM [blood_final]
JOIN [test_data] 
ON [test_data].[test_id]=[blood_final].[test_id] 
AND [test_data].[smear_id]=[blood_final].[smear_id] 
AND [test_data].[count]=[blood_final].[count]
JOIN [bloodtest].[dbo].[id] ON [id].[No] = [test_data].[test_id]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id] = [test_data].[smear_id]
WHERE [id].[院區]='%s' AND [bloodinfo].[year]=%d;"""%(h_site,int(testyear))
                file_name = "%d_%s_"%(int(testyear),h_site)
            else:
                tk.messagebox.showerror(title='土城醫院檢驗科', message='請選擇年份!')
                return
        else:
            tk.messagebox.showerror(title='土城醫院檢驗科', message='請選擇院區!所有院區請選擇ALL')
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet1"

        title = ["院區","姓名","年份","考片ID","count","celltype","matrix_value","must","mustnot","abn_lym","score"]
        ws.append(title)
        with engine.begin() as conn:
            sas_data = pd.read_sql_query(sa.text(get_sas), conn)
        for i in range(0,len(sas_data)):
            ws.append(sas_data.iloc[i].to_list())

        wb.save(file_name + "SAS.xlsx")
        filepath = ".//%sSAS.xlsx"%(file_name)
        if os.path.isfile(filepath):
            tk.messagebox.showinfo(title='土城醫院檢驗科', message='檔案新增成功!')
        else:
            tk.messagebox.showerror(title='土城醫院檢驗科', message='檔案新增失敗QQ')
##################↑↑↑↑↑匯出SAS↑↑↑↑↑##################
##清除按鈕
    def clear(self):
        ##清除兩個combobox
        self.all_combobox.set("")
        self.year_combobox.set("")
        ##清除一個tksheet內容
        self.sht_result.set_cell_data()
        ##鎖定三個匯出按鈕
        self.btn_export.configure(state='disabled')
        self.btn_rawexport.configure(state='disabled')
        self.btn_sasexport.configure(state='disabled')

##返回basdesk按鈕    
    def back(self,oldmaster):
        try:
            oldmaster.deiconify()
            self.master.withdraw()
        except AttributeError:
            self.master.destroy()
            return


def main():  
    root = ctk.CTk()
    OP = OUTPUT_SCORE(root)
    # root['bg']='#FFEEDD'
    # B.gui_arrang()
    # 主程式執行  
    root.mainloop()  
  
  
if __name__ == '__main__':  
    main()  