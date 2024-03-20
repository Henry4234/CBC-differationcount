import openpyxl,pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from sqlalchemy.engine import URL
from sqlalchemy import create_engine
import sqlalchemy as sa
import pandas as pd


connection_string = """DRIVER={ODBC Driver 17 for SQL Server};SERVER=localhost;DATABASE=bloodtest;UID=sa;PWD=1234"""
try:
    coxn = pyodbc.connect(connection_string)
except pyodbc.OperationalError:
    connection_string = "DRIVER={ODBC Driver 17 for SQL Server};SERVER=10.30.47.9;DATABASE=bloodtest;UID=henry423;PWD=1234"
finally:
    coxn = pyodbc.connect(connection_string)
connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": connection_string})
engine = create_engine(connection_url)

# with coxn.cursor() as cursor:
#     # query = 'SELECT "year","smear_id" FROM [bloodtest].[dbo].[bloodinfo];'
#     query = "SELECT [bloodinfo_ans2].[smear_id],[bloodinfo_ans2].[celltype],[bloodinfo_ans2].[value] FROM [bloodtest].[dbo].[bloodinfo_ans2];"
#     cursor.execute(query)
#     db2 = dict(cursor.fetchall())
# print(db2)

with engine.begin() as conn:
    filter_test = """SELECT [bloodinfo_ans2].[smear_id],[bloodinfo].[count_value],[bloodinfo_ans2].[celltype],[bloodinfo_ans2].[value]
FROM [bloodtest].[dbo].[bloodinfo_ans2]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id] = [bloodinfo_ans2].[smear_id]
WHERE [bloodinfo_ans2].[smear_id] IN (SELECT DISTINCT [blood_final].[smear_id] FROM [bloodtest].[dbo].[blood_final]);"""
    datadict2 = pd.read_sql_query(sa.text(filter_test), conn)
    lst = datadict2['smear_id'].unique()
# print(lst)
# print(datadict2)

# datadict2 = datadict2.to_dict()
# print(datadict2)
# print(lst)




wb = Workbook()
ws = wb.active
ws.title = "demo"

title = ["考片ID","count_value","plasma cell","abnormal lympho","megakaryocyte","nRBC","blast","metamyelocyte","eosinophil","plasmacytoid","promonocyte","promyelocyte","band neutropil","basopil","atypical lymphocyte","hypersegmented neutrophil","myelocyte","segmented neutrophil","lymphocyte","monocyte"]

ws.append(title)
n=2
for i in range(0,len(lst)):
    flt = (datadict2['smear_id']==lst[i])
    flt_dic = datadict2.loc[flt]
    dic_2 = flt_dic.set_index('celltype')['value'].to_dict()
    # ws.append([item])
    # print(dic_2)
    dic_2['smear_id'] = lst[i]
    dic_2['count_value'] = flt_dic['count_value'].iloc[0]

    ws.cell(row=n, column=1, value=lst[i])
    ws.cell(row=n, column=2, value=flt_dic['count_value'].iloc[0])
    # 逐列遍历 row(1)，查找对应的值并放入 row(n)
    for col in range(3, 21):  # 从第二列开始遍历
        celltype = ws.cell(row=1, column=col).value
        if celltype in dic_2:
            value = dic_2[celltype]
            ws.cell(row=n, column=col, value=value)
    n+=1
#空排兩排
# ws.insert_rows(len(lst)+2,2)
title_2=['院區','姓名','year','考片ID','count','celltype','percent_value','timestamp','lower','upper']
for i in range(0,len(title_2)):
    char = get_column_letter(i+1)
    ws[char + str(len(lst)+4)] = title_2[i]

with engine.begin() as conn:
    filter_final = """SELECT [id].[院區],[id].[ac],[bloodinfo].[year],[blood_final].[smear_id],[blood_final].[count],[blood_final].[celltype],[blood_final].[percent_value],[blood_final].[timestamp]
FROM [bloodtest].[dbo].[blood_final]
JOIN [bloodtest].[dbo].[id] ON [id].[No] = [blood_final].[test_id]
JOIN [bloodtest].[dbo].[bloodinfo] ON [bloodinfo].[smear_id] = [blood_final].[smear_id];"""
    data = pd.read_sql_query(sa.text(filter_test), conn)
# print(data)
#加入data
for j in range(0,len(data)):
    ws.append(data.iloc[j].to_list())
    ws["I" + str(len(lst)+5+j)] = "=VLOOKUP(ROUND(G%s,0),'95'!$A$1:$K$103,IFS(VLOOKUP(D%s,$A$2:$B$%s,2,FALSE)=100,2,VLOOKUP(D%s,$A$2:$B$%s,2,FALSE)=200,4,VLOOKUP(D%s,$A$2:$B$%s,2,FALSE)=500,6),FALSE)"%(str(len(lst)+5+j),str(len(lst)+5+j),str(len(lst)+1),str(len(lst)+5+j),str(len(lst)+1),str(len(lst)+5+j),str(len(lst)+1))
    ws["J" + str(len(lst)+5+j)] = "=VLOOKUP(ROUND(G%s,0),'95'!$A$1:$K$103,IFS(VLOOKUP(D%s,$A$2:$B$%s,2,FALSE)=100,3,VLOOKUP(D%s,$A$2:$B$%s,2,FALSE)=200,5,VLOOKUP(D%s,$A$2:$B$%s,2,FALSE)=500,7),FALSE)"%(str(len(lst)+5+j),str(len(lst)+5+j),str(len(lst)+1),str(len(lst)+5+j),str(len(lst)+1),str(len(lst)+5+j),str(len(lst)+1))
wb.save("test_01.xlsx")